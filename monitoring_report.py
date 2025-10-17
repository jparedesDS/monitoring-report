import time
import warnings
import pandas as pd
from openpyxl import load_workbook
warnings.filterwarnings('ignore')
start_time = time.time()

# === RUTA Y LECTURA DE DATOS ===
data_path_total = r'C:\Users\alejandro.berzal\Desktop\DATA SCIENCE\new-monitoring-report\data_import\data_erp.xlsx'
erp_data = pd.read_excel(data_path_total)
df_total = pd.read_excel(data_path_total)
consultar_data_path_total = r'C:\Users\alejandro.berzal\Desktop\DATA SCIENCE\new-monitoring-report\data_import\consulta_erp.xlsx'
consulta_data = pd.read_excel(consultar_data_path_total)

# === UNIR COLUMNAS 'Responsable' Y 'Nº Oferta' DESDE consulta_data ===
cols_to_add = ['Nº Pedido', 'Responsable', 'Nº Oferta']
missing_cols = [col for col in cols_to_add if col not in consulta_data.columns]
if missing_cols:
    print(f"⚠️ Columnas faltantes en consulta_data: {missing_cols}")
else:
    df_total = df_total.merge(consulta_data[cols_to_add], on='Nº Pedido', how='left')
    erp_data = erp_data.merge(consulta_data[cols_to_add], on='Nº Pedido', how='left')

# === LIMPIEZA Y FORMATEO DE FECHAS ===
erp_data['Estado'] = erp_data['Estado'].fillna('Sin Enviar')
df_total['Estado'] = df_total['Estado'].fillna('Sin Enviar')
erp_data = erp_data[erp_data['Estado'] != 'Eliminado'].copy()
df_total = df_total[df_total['Estado'] != 'Eliminado'].copy()
today_date = pd.to_datetime('today')
today_date_str = today_date.strftime('%d-%m-%Y')

for df in [erp_data, df_total, consulta_data]:
    for col in ['Fecha', 'Fecha Pedido', 'Fecha Prevista', 'Fecha Fabricación', 'Fecha Montaje', 'Fecha Envío']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True)

# === DATAFRAMES SEGÚN ESTADO ===
df_comentados = erp_data[erp_data['Estado'].isin(['Com. Menores', 'Com. Mayores', 'Rechazado', 'Comentado'])].copy()
df_envio = erp_data[erp_data['Estado'] == 'Enviado'].copy()
df_sin_envio = erp_data[erp_data['Estado'] == 'Sin Enviar'].copy()
df_criticos = erp_data[(erp_data['Crítico'] == 'Sí') & (~erp_data['Estado'].isin(['Eliminado', 'Aprobado', 'Enviado']))].copy()

# === COLUMNAS CALCULADAS (Días Devolución) ===
for df, fecha_ref in [(df_comentados, 'Fecha'), (df_envio, 'Fecha'), (df_criticos, 'Fecha')]:
    if fecha_ref in df.columns:
        df['Días Devolución'] = (today_date - df[fecha_ref]).dt.days
if 'Fecha Pedido' in df_sin_envio.columns:
    df_sin_envio['Días Devolución'] = (today_date - df_sin_envio['Fecha Pedido']).dt.days

# === RENOMBRAR COLUMNAS CLAVE ===
def rename_columns(df, fecha_old, new_name):
    if fecha_old in df.columns:
        df = df.rename(columns={fecha_old: new_name})
    return df
df_comentados = rename_columns(df_comentados, 'Fecha', 'Fecha Dev. Doc.')
df_envio = rename_columns(df_envio, 'Fecha', 'Fecha Env. Doc.')
df_criticos = rename_columns(df_criticos, 'Fecha', 'Fecha Doc.')
df_total = rename_columns(df_total, 'Fecha', 'Fecha Doc.')
df_sin_envio = rename_columns(df_sin_envio, 'Fecha', 'Fecha Doc.')

# === CÁLCULO DE NOTAS EN df_comentados ===
df_comentados.insert(12, "Notas", df_comentados['Fecha Dev. Doc.'])
mask = df_comentados['Estado'].isin(['Rechazado', 'Com. Menores', 'Com. Mayores', 'Comentado'])
df_comentados.loc[mask, 'Notas'] += pd.to_timedelta(15, unit='D')
df_comentados['Notas'] = "Enviar antes del " + df_comentados['Notas'].dt.strftime('%d-%m-%Y')

# === STATUS GLOBAL ===
status_global = (
    erp_data.groupby(['Nº Pedido', 'Estado'])
    .size()
    .unstack(fill_value=0)
    .reset_index()
)
if 'Eliminado' in status_global.columns:
    status_global = status_global.drop(columns='Eliminado')
for col in ['Aprobado', 'Com. Mayores', 'Com. Menores', 'Enviado', 'Rechazado', 'Sin Enviar']:
    if col not in status_global.columns:
        status_global[col] = 0
status_global['Total'] = status_global.iloc[:, 1:].sum(axis=1)
status_global['% Completado'] = (status_global['Aprobado'] / status_global['Total'] * 100).fillna(0).round(2)
status_global = status_global[status_global['% Completado'] != 100]

# === UNIFICAR PENDIENTES ===
#df_pendientes = pd.concat([df_sin_envio, df_comentados], ignore_index=True)

# === ORDENAR COLUMNAS ===
column_order = [
    'Nº Pedido', 'Responsable', 'Nº Oferta', 'Nº PO', 'Cliente', 'Material',
    'Fecha Pedido', 'Fecha Prevista', 'Nº Doc. Cliente', 'Nº Doc. EIPSA',
    'Título', 'Tipo Doc.', 'Info/Review', 'Repsonsable', 'Días Envío', 'Crítico', 'Estado', 'Notas', 'Nº Revisión',
    'Fecha Doc.', 'Fecha Env. Doc.', 'Fecha Dev. Doc.', 'Días Devolución',
    'Reclamaciones', 'Seguimiento', 'Historial Rev.'
]
def reorder_columns(df):
    existing_cols = [col for col in column_order if col in df.columns]
    remaining_cols = [col for col in df.columns if col not in existing_cols]
    return df[existing_cols + remaining_cols]
df_total = reorder_columns(df_total)
df_envio = reorder_columns(df_envio)
df_comentados = reorder_columns(df_comentados)
df_criticos = reorder_columns(df_criticos)
df_sin_envio = reorder_columns(df_sin_envio)

# === EXPORTAR A EXCEL ===
output_path = fr'C:\Users\alejandro.berzal\Desktop\DATA SCIENCE\new-monitoring-report\monitoring_report_{today_date_str}.xlsx'
with pd.ExcelWriter(output_path, engine='openpyxl', datetime_format='DD/MM/YYYY') as writer:
    # Orden de las sheets:
    # ENVIADOS
    df_envio.to_excel(writer, sheet_name='ENVIADOS', index=False)
    # PENDIENTES
    df_comentados.to_excel(writer, sheet_name='DEVOLUCIONES', index=False)
    # CRÍTICOS
    df_criticos_menor15 = df_criticos[(df_criticos['Días Devolución'] <= 15) | (df_criticos['Días Devolución'].isna())].copy()
    df_criticos_menor15.to_excel(writer, sheet_name='CRÍTICOS', index=False)
    # CRÍTICOS +15d
    df_criticos_mas15 = df_criticos[df_criticos['Días Devolución'] > 15].copy()
    df_criticos_mas15.to_excel(writer, sheet_name='CRÍTICOS +15d', index=False)
    # SIN ENVIAR
    df_sin_envio.to_excel(writer, sheet_name='SIN ENVIAR', index=False)
    # ALL DOC.
    df_total.to_excel(writer, sheet_name='ALL DOC.', index=False)
    # GRÁFICO EN "STATUS GLOBAL"
    status_global.to_excel(writer, sheet_name='STATUS GLOBAL', index=False)

# === FORMATO FECHAS, FILTRO, ORDEN Y AJUSTE COLUMNAS ===
fechas_cols = ["Fecha", "Fecha Pedido", "Fecha Prevista", "Fecha Dev. Doc.", "Fecha Env. Doc.", "Fecha Doc."]
wb = load_workbook(output_path)
for ws in wb.worksheets:
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    # Formato fechas
    for col_name in fechas_cols:
        if col_name in header_map:
            col_idx = header_map[col_name]
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cell:
                    c.number_format = 'DD/MM/YYYY'

    # Filtro automático
    ws.auto_filter.ref = ws.dimensions

    # Orden descendente por Nº Pedido
    if 'Nº Pedido' in header_map:
        col_idx = header_map['Nº Pedido']
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows_sorted = sorted(data_rows, key=lambda x: x[col_idx - 1] if x[col_idx - 1] else 0, reverse=True)
        for i, row in enumerate(data_rows_sorted, start=2):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)

    # Ajuste automático ancho columnas según encabezado
    for col_cells in ws.iter_cols(min_row=1, max_row=1):
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

wb.save(output_path)


# === GRÁFICO DE BARRAS APILADAS EN "STATUS GLOBAL" ===
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties

wb = load_workbook(output_path)

if "STATUS GLOBAL" in wb.sheetnames:
    ws = wb["STATUS GLOBAL"]

    # Buscar encabezados
    headers = [cell.value for cell in ws[1]]

    # Columnas a graficar
    estado_cols = ["Aprobado", "Com. Mayores", "Com. Menores", "Enviado", "Rechazado", "Sin Enviar"]
    col_indices = [headers.index(col) + 1 for col in estado_cols]  # +1 porque Excel es 1-based

    # Rango de pedidos (columna A)
    min_row = 2
    max_row = ws.max_row
    pedidos = Reference(ws, min_col=1, min_row=min_row, max_row=max_row)

    # Rango de datos
    min_col = min(col_indices)
    max_col = max(col_indices)
    data = Reference(ws, min_col=min_col, max_col=max_col, min_row=1, max_row=max_row)

    # Crear gráfico de barras apiladas
    chart = BarChart()
    chart.type = "col"
    chart.title = "Estado por Pedido"
    chart.style = 12
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.y_axis.title = "Nº Documentos"
    chart.x_axis.title = "Nº Pedido"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(pedidos)

    # Colores armonizados exactos para cada estado
    colores = ["00B350",  # Aprobado
               "C59B3F",  # Com. Mayores
               "FFCF7F",  # Com. Menores
               "5566A0",  # Enviado
               "FF8273",  # Rechazado
               "FFEF7F"]  # Sin Enviar

    for i, serie in enumerate(chart.series):
        serie.graphicalProperties = GraphicalProperties(solidFill=colores[i])

    chart.height = 16
    chart.width = 29

    # Insertar gráfico en la hoja
    ws.add_chart(chart, "J3")

    wb.save(output_path)
    print("✅ Gráfico de barras apiladas añadido correctamente en la hoja de datos: 'STATUS GLOBAL'.")
else:
    print("⚠️ ERROR: No se encontró la hoja 'STATUS GLOBAL'.")


# === FUNCIONES DE COLORES Y ESTILOS EXCEL ===
def apply_excel_styles(archivo_excel):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, NamedStyle
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import Font

    workbook = load_workbook(archivo_excel)

    # === Estilos básicos ===
    fill_light = PatternFill(start_color="D4DCF4", end_color="D4DCF4", fill_type="solid")
    fill_dark = PatternFill(start_color="6678AF", end_color="6678AF", fill_type="solid")
    font_white = Font(color='FFFFFF', bold=True)
    font_black = Font(color='000000')
    border_medium = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

    # === Regla fila Días Devolución > 15 en PENDIENTES ===
    diff_pendientes = DifferentialStyle(fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))

    # === Función para aplicar estilo general a una hoja ===
    def style_sheet(sheet, tab_color):
        sheet.sheet_properties.tabColor = tab_color
        sheet.freeze_panes = sheet['B2']
        max_col = sheet.max_column
        max_row = sheet.max_row

        # Cabecera
        for cell in sheet[1]:
            cell.fill = fill_dark
            cell.font = font_white
            cell.border = border_medium

        # Celdas del resto del cuerpo
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = fill_light
                cell.font = font_black
                cell.border = border_medium

        # Filtro automático
        sheet.auto_filter.ref = sheet.dimensions

        # Regla fila completa PENDIENTES Días Devolución > 15
        if sheet.title == "DEVOLUCIONES":
            col_idx = None
            for idx, cell in enumerate(sheet[1], start=1):
                if cell.value == "Días Devolución":
                    col_idx = idx
                    break
            if col_idx:
                col_letter = sheet.cell(row=2, column=col_idx).column_letter
                formula = f"=${col_letter}2>15"
                rule_full_row = Rule(type="expression", formula=[formula], dxf=diff_pendientes, stopIfTrue=False)
                rango_filas = f"A2:{sheet.cell(row=2, column=max_col).column_letter}{max_row}"
                sheet.conditional_formatting.add(rango_filas, rule_full_row)

    # === Colores de pestañas ===
    hoja_colores = {
        'ALL DOC.': "6678AF",
        'ENVIADOS': "00D25F",
        'DEVOLUCIONES': "FFA19A",
        'CRÍTICOS': "DBB054",
        'CRÍTICOS +15d': "FF7F50",
        'SIN ENVIAR' : "FFFF66",
        'STATUS GLOBAL': "B1E1B9"
    }

    # Aplicar estilos a todas las hojas
    for hoja, color in hoja_colores.items():
        if hoja in workbook.sheetnames:
            style_sheet(workbook[hoja], color)

    # === Colorear columna Estado con fondo y negrita ===
    estado_colores = {
        "Rechazado": "FFA19A",
        "Com. Menores": "FFE5AD",
        "Com. Mayores": "DBB054",
        "Comentado": "F79646",
        "Enviado": "B1E1B9",
        "Sin Enviar": "FFFFAB",
        "Información": "FFFF46",
        "HOLD": "FF0909",
        "Aprobado": "00D25F"
    }
    # Aplicar estilos a todas las hojas
    for sheet in workbook.worksheets:
        col_estado_idx = None
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value == "Estado":
                col_estado_idx = idx
                break
        if col_estado_idx:
            for row in sheet.iter_rows(min_row=2):
                estado_val = row[col_estado_idx - 1].value
                if estado_val in estado_colores:
                    color_hex = estado_colores[estado_val]
                    row[col_estado_idx - 1].fill = PatternFill(start_color=color_hex,
                                                               end_color=color_hex,
                                                               fill_type="solid")
                    #row[col_estado_idx - 1].font = Font(bold=True) #Añadir que sea negrita


    # === Colorear columnas Responsable y Repsonsable solo texto y negrita ===
    responsables_colores = {
        "SS": "A32121",  # Azul marino oscuro
        "JM": "2C8A6A",  # Verde esmeralda
        "JV": "006B95",  # Azul petróleo
        "EC": "B31274",  # Rosa intenso (magenta oscuro)
        "ES": "5A0DA0",  # Púrpura fuerte
        "JP": "00458F",  # Azul profesional
        "AC": "3F0075",  # Índigo
        "CCH": "1F1F1F",  # Gris oscuro casi negro
        "LB": "176DD1",  # Azul brillante
        "RM": "228B22",  # Verde bosque
        "RP": "1B365D",  # Rojo oscuro
        "EC/SS": "1F7A1F" # Verde bosque
    }

    for sheet in workbook.worksheets:
        # Buscar índices de las columnas 'Responsable' y 'Repsonsable'
        col_indices = []
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value in ["Responsable", "Repsonsable"]:
                col_indices.append(idx)

        # Aplicar colores y negrita a ambas columnas si existen
        for col_idx in col_indices:
            for row in sheet.iter_rows(min_row=2):
                val = row[col_idx - 1].value
                if val in responsables_colores:
                    color_hex = responsables_colores[val]
                    row[col_idx - 1].font = Font(color=color_hex, bold=True)

    # === Colorear datos en las columnas indicadas (colores, negrita y reglas condicionales) ===
    for sheet in workbook.worksheets:
        # Buscar índices de las columnas
        col_estado_idx = None
        col_critico_idx = None
        col_dias_idx = None
        col_info_rev_idx = None
        col_dias_env_idx = None
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value == "Nº Pedido":
                col_pedido_idx = idx
            if cell.value == "Estado":
                col_estado_idx = idx
            if cell.value == "Crítico":
                col_critico_idx = idx
            if cell.value == "Info/Review":
                col_info_rev_idx = idx
            if cell.value == "Días Envío":
                col_dias_env_idx = idx
            elif cell.value == "Días Devolución":
                col_dias_idx = idx

        # Aplicar formato independiente por columna
        if col_estado_idx:
            for row in sheet.iter_rows(min_row=2):
                if row[col_estado_idx - 1].value == "HOLD":
                    row[col_estado_idx - 1].font = Font(color="FF0000", bold=True)
                elif row[col_estado_idx - 1].value == "Aprobado":
                    row[col_estado_idx - 1].font = Font(bold=True)

        if col_critico_idx:
            for row in sheet.iter_rows(min_row=2):
                if row[col_critico_idx - 1].value == "Sí":
                    row[col_critico_idx - 1].font = Font(color="FF0000", bold=True)
                #elif row[col_critico_idx - 1].value == "No":
                    #row[col_critico_idx - 1].font = Font(color="000000", bold=True)

        if col_info_rev_idx:
            for row in sheet.iter_rows(min_row=2):
                if row[col_info_rev_idx - 1].value == "R":
                    row[col_info_rev_idx - 1].font = Font(color="FF0000", bold=True)
                elif row[col_info_rev_idx - 1].value == "I":
                    row[col_info_rev_idx - 1].font = Font(color="4D4D4D", bold=True)

        if col_dias_env_idx:
            for row in sheet.iter_rows(min_row=2):
                if row[col_dias_env_idx - 1].value == 15:
                    row[col_dias_env_idx - 1].font = Font(color="4D4D4D", bold=True)

        if col_dias_idx:
            for row in sheet.iter_rows(min_row=2):
                valor = row[col_dias_idx - 1].value
                if valor is not None and isinstance(valor, (int, float)) and valor > 15:
                    row[col_dias_idx - 1].font = Font(color="FF0000", bold=True)

        '''# Poner toda la primera columna en negrita
        for row in sheet.iter_rows(min_row=1):
            cell = row[0]  # columna A
            # Mantener bold si ya estaba en color, pero garantizar negrita
            if cell.font:
                cell.font = Font(name=cell.font.name, size=cell.font.size, bold=True, color=cell.font.color)
            else:
                cell.font = Font(bold=True)'''

    # === Ajustar ancho de todas las columnas automáticamente ===
    for sheet in workbook.worksheets:
        for col_cells in sheet.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            sheet.column_dimensions[col_letter].width = max_length + 2

        # === Centrar columnas específicas ===
    from openpyxl.styles import Alignment
    columnas_a_centrar = ['Responsable', 'Repsonsable', 'Info/Review', 'Días Envío', 'Crítico', 'Estado', 'Días Devolución', 'Nº Revisión',
                          'Aprobado', 'Com. Mayores', 'Com. Menores', 'Enviado', 'Rechazado', 'Sin Enviar', 'Total', '% Completado']

    for sheet in workbook.worksheets:
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value in columnas_a_centrar:
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                    row[idx - 1].alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(archivo_excel)
    print("✅ Estilos aplicados a toda la tabla de datos.")

# === APLICAR ESTILOS ===
apply_excel_styles(output_path)

print(f"✅ Archivo Excel final generado en:\n{output_path}")
print("Duración del proceso: {:.2f} segundos".format(time.time() - start_time))