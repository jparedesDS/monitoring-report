# Imports
from openpyxl import load_workbook
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Font
from openpyxl.chart import BarChart, Reference
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from datetime import datetime

# Función aplicable para el tratamiento y coloreado de datos en la tabla de excel
def highlight_row_content(column, value, color):
    cont_val = column == value
    return [f'background-color: {color}' if cont_val.any() else '' for v in cont_val]


# Función para definir el rango de las celdas
def auto_fit_columns(sheet):
    if sheet:
        for col_index in range(sheet.getCells().getMaxDataColumn() + 1):
            sheet.autoFitColumn(col_index)


# Coloreado y estilos de la tabla de excel
def apply_excel_styles(today_date):
    # Carga el archivo de Excel existente
    archivo_excel = "Monitoring_Report_" + str(today_date) + ".xlsx"
    workbook = load_workbook(archivo_excel)

    # Definir los estilos
    fecha_style = NamedStyle(name='fecha')
    fecha_style.number_format = 'DD-MM-YYYY'
    cell_filling_blue_light = PatternFill(start_color="D4DCF4", end_color="D4DCF4", fill_type="solid")
    cell_filling = PatternFill(start_color="6678AF", end_color="6678AF", fill_type="solid")
    cell_filling_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    medium_dashed = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    font_white = Font(color='FFFFFF', bold=True)
    font_black = Font(color='000000')

    # Definir el estilo condicional para fechas de pedido superiores a las fechas previstas
    red_fill = PatternFill(start_color="FF5B5B", end_color="FF5B5B", fill_type="solid")
    diff_style = DifferentialStyle(fill=red_fill)
    rule = Rule(type="expression", dxf=diff_style)
    rule.formula = ["$A2>$B2"]  # Ajustar según las columnas de fechas

    # Definir el estilo condicional para días de devolución >= 15
    red_fill_2 = Font(color="FF5B5B", bold=True)
    diff_style_devolucion = DifferentialStyle(font=red_fill_2)
    rule_devolucion = Rule(type="cellIs", operator="greaterThanOrEqual", formula=["15"], dxf=diff_style_devolucion)

    # Función para aplicar estilos a una hoja
    def apply_styles_to_sheet(sheet, tab_color, max_row, max_col, column_exceptions=('K',)):
        freeze = sheet['B2']
        sheet.freeze_panes = freeze
        sheet.sheet_properties.tabColor = tab_color

        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            if row_idx == 1:
                for cell in row:
                    if cell.column_letter not in column_exceptions:
                        cell.style = fecha_style
                        cell.fill = cell_filling
            else:
                for cell in row:
                    if cell.column_letter not in column_exceptions:
                        if isinstance(cell.value, datetime):
                            cell.style = fecha_style
                        cell.fill = cell_filling_blue_light

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = medium_dashed

        cell_letters = sheet['K1']
        cell_letters.fill = cell_filling
        cell_letters = sheet['L1']
        cell_letters.fill = cell_filling

        for cell in sheet[1]:
            cell.font = font_white

        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            if row_idx != 1:
                for cell in row:
                    cell.font = font_black

        for fila in sheet.iter_rows(min_row=1, max_row=500, min_col=1, max_col=max_col):
            for celda in fila:
                if celda.value == 'Sí':
                    celda.font = Font(color='FF5B5B', bold=True)
                if celda.value == 'LB':
                    celda.font = Font(color='0072C8', bold=True)
                if celda.value == 'AC':
                    celda.font = Font(color='7030A0', bold=True)
                if celda.value == 'SS':
                    celda.font = Font(color='FF5B5B', bold=True)
                if celda.value == 100:
                    celda.font = Font(color='FF5B5B', bold=True)

        sheet.auto_filter.ref = f"A1:{chr(65 + max_col - 1)}{max_row}"

        # Aplicar la regla de formato condicional fechas pedidos y prevista
        #sheet.conditional_formatting.add(f"P2:Q2{max_row}", rule)
        # Definir el estilo condicional para días de devolución >= 15
        sheet.conditional_formatting.add(f"O2:O{max_row}", rule_devolucion)


    def add_chart(sheet):
        chart = BarChart()
        chart.type = "col"  # Configurar el gráfico como gráfico de columnas
        chart.title = "Estado de la Documentación (PENDIENTES)"
        chart.y_axis.title = 'PORCENTAJE COMPLETADO'
        chart.x_axis.title = 'Nº DE PEDIDOS'
        chart.style = 10
        chart.varyColors = "0000FFFF"

        filtered_categories = []
        filtered_data = []
        for row in range(2, sheet.max_row + 1):  # Asumiendo que la primera fila tiene encabezados
            porcentaje_completado = sheet.cell(row=row, column=2).value
            if porcentaje_completado is not None and porcentaje_completado < 100:
                pedido = sheet.cell(row=row, column=1).value
                filtered_categories.append(pedido)
                filtered_data.append(porcentaje_completado)
        # Escribimos las categorías y datos filtrados en columnas temporales al final de la hoja
        start_col = sheet.max_column + 20
        for idx, value in enumerate(filtered_categories, start=2):
            sheet.cell(row=idx, column=start_col, value=value)
        for idx, value in enumerate(filtered_data, start=2):
            sheet.cell(row=idx, column=start_col + 1, value=value)

        data = Reference(sheet, min_col=start_col + 1, min_row=1, max_row=len(filtered_data) + 1)
        categories = Reference(sheet, min_col=start_col, min_row=2, max_row=len(filtered_categories) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.smooth = True
        # Posicionar y escalar el gráfico
        sheet.add_chart(chart, "L2")
        chart.width = 27  # Ancho del gráfico en pulgadas
        chart.height = 17  # Alto del gráfico en pulgadas

        for col in ['K', 'L']:
            cell = sheet[f'{col}1']
            cell.fill = PatternFill(fill_type=None)

    def add_stacked_bar_chart_cal_pla(sheet):
        chart = BarChart()
        chart.type = "col"
        chart.title = "Estado Doc. Cálculos y Planos (ESP-0003)"
        chart.style = 10
        chart.y_axis.title = 'DOCUMENTOS'
        chart.x_axis.title = 'Nº DE PEDIDO'
        chart.varyColors = "0000FFFF"

        # Encontrar la fila que contiene los datos
        min_row = None
        for row in sheet.iter_rows():
            if min_row is None:
                for cell in row:
                    if cell.value is not None:
                        min_row = row[2].row
                        break

        # Si no se encontraron datos, salir de la función
        if min_row is None:
            print("No se encontraron datos en la hoja.")
            return

        max_row = sheet.max_row
        max_col = 3

        # Determinar el rango de columnas
        min_col = 2  # Empezamos desde la segunda columna (suponiendo que la primera contiene etiquetas)
        for cell in sheet[min_row]:
            if cell.value is not None:
                break
            min_col += 1

        # Seleccionar los datos y categorías
        data = Reference(sheet, min_col=min_col, min_row=min_row, max_row=max_row, max_col=max_col)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=max_row, max_col=1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Habilitar etiquetas de datos
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        # Posicionar y escalar el gráfico
        sheet.add_chart(chart, "E1")
        chart.width = 15
        chart.height = 10

        for col in ['K', 'L', 'J']:
            cell = sheet[f'{col}1']
            cell.fill = PatternFill(fill_type=None)

    def add_stacked_bar_chart_planos(sheet):
        chart = BarChart()
        chart.type = "col"
        chart.title = "Estado Doc. Planos (PLG-0005)"
        chart.style = 10
        chart.y_axis.title = 'DOCUMENTOS'
        chart.x_axis.title = 'Nº DE PEDIDO'
        chart.varyColors = "0000FFFF"

        # Encontrar la fila que contiene los datos
        min_row = None
        for row in sheet.iter_rows():
            if min_row is None:
                for cell in row:
                    if cell.value is not None:
                        min_row = row[2].row
                        break

        # Si no se encontraron datos, salir de la función
        if min_row is None:
            print("No se encontraron datos en la hoja.")
            return

        max_row = sheet.max_row
        max_col = 14

        # Determinar el rango de columnas
        min_col = 13  # Empezamos desde la segunda columna (suponiendo que la primera contiene etiquetas)
        for cell in sheet[min_row]:
            if cell.value is not None:
                break
            min_col += 1

        # Seleccionar los datos y categorías
        data = Reference(sheet, min_col=min_col, min_row=min_row, max_row=max_row, max_col=max_col)
        categories = Reference(sheet, min_col=12, min_row=2, max_row=max_row, max_col=12)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Habilitar etiquetas de datos
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        # Posicionar y escalar el gráfico
        sheet.add_chart(chart, "P1")
        chart.width = 15
        chart.height = 10

        # Iterar a través de cada columna y fila para aplicar los cambios
        start_row = 1
        end_row = 10
        for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K',]:
            for row in range(start_row, end_row + 1):
                cell = sheet[f'{col}{row}']
                cell.fill = PatternFill(fill_type=None)  # Eliminar el relleno
                cell.border = Border()  # Eliminar los bordes
        cell_letters = sheet['L1']
        cell_letters.fill = cell_filling

    # Aplicar estilos a cada hoja
    apply_styles_to_sheet(workbook['DOC. COMENTARIOS'], "DBB054", 200, 21, ('K', 'L'))
    apply_styles_to_sheet(workbook['DOC. ENVIADA'], "B1E1B9", 200, 20, ('K'))
    apply_styles_to_sheet(workbook['DOC. SIN ENVIAR'], "DDDDDD", 200, 13, ('K'))
    apply_styles_to_sheet(workbook['CRÍTICOS'], "FFFF46", 200, 13, ('K'))
    grafico_sheet = workbook['ESTADO GLOBAL']
    apply_styles_to_sheet(workbook['DOC. TOTAL'], "99CCFF", 500, 18, ('K'))
    apply_styles_to_sheet(grafico_sheet, "FFAAAB", 110, 10, ('K','L','M'))
    add_chart(grafico_sheet)
    grafico_planos = workbook['GRÁFICO CRÍTICOS']
    apply_styles_to_sheet(grafico_planos, "FFFFAB", 2, 3)
    add_stacked_bar_chart_cal_pla(grafico_planos)
    grafico_planos = workbook['GRÁFICO CRÍTICOS']
    add_stacked_bar_chart_planos(grafico_planos)
    # Guardar el archivo modificado
    workbook.save(archivo_excel)
    print("¡Creando los filtros de las columnas!")

