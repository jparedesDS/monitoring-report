# === FUNCIONES DE COLORES Y ESTILOS EXCEL ===
def apply_excel_styles(archivo_excel):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, NamedStyle
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import Font

    workbook = load_workbook(archivo_excel)

    # === Estilos básicos ===
    fill_light = PatternFill(start_color="D4DCF4", end_color="D4DCF4", fill_type="solid")
    fill_dark = PatternFill(start_color="6678AF", end_color="6678AF", fill_type="solid")
    font_white = Font(color='FFFFFF', bold=True)
    font_black = Font(color='000000')
    border_medium = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

    # === Regla fila Días Devolución > 15 en PENDIENTES ===
    diff_pendientes = DifferentialStyle(fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))

    # === Función para aplicar estilo general a una hoja ===
    def style_sheet(sheet, tab_color):
        sheet.sheet_properties.tabColor = tab_color
        sheet.freeze_panes = sheet['B2']
        max_col = sheet.max_column
        max_row = sheet.max_row

        # Cabecera
        for cell in sheet[1]:
            cell.fill = fill_dark
            cell.font = font_white
            cell.border = border_medium

        # Celdas del resto del cuerpo
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = fill_light
                cell.font = font_black
                cell.border = border_medium

        # === Filtro automático ===
        sheet.auto_filter.ref = sheet.dimensions

        # === Regla fila completa PENDIENTES Días Devolución > 15 ===
        if sheet.title == "DEVOLUCIONES":
            col_idx = None
            for idx, cell in enumerate(sheet[1], start=1):
                if cell.value == "Días Devolución":
                    col_idx = idx
                    break
            if col_idx:
                col_letter = sheet.cell(row=2, column=col_idx).column_letter
                formula = f"=${col_letter}2>15"
                rule_full_row = Rule(type="expression", formula=[formula], dxf=diff_pendientes, stopIfTrue=False)
                rango_filas = f"A2:{sheet.cell(row=2, column=max_col).column_letter}{max_row}"
                sheet.conditional_formatting.add(rango_filas, rule_full_row)

    # === Colores de pestañas ===
    hoja_colores = {
        'ALL DOC.': "0072C8",
        'ENVIADOS': "B1E1B9",
        'DEVOLUCIONES': "E26B0A",
        'CRÍTICOS': "FFFF46",
        'CRÍTICOS +15d': "FF5B5B",
        'SIN ENVIAR' : "FFFFAB",
        'STATUS GLOBAL': "FFAAAB"
    }

    # === Aplicar estilos a todas las hojas ===
    for hoja, color in hoja_colores.items():
        if hoja in workbook.sheetnames:
            style_sheet(workbook[hoja], color)

    # === Colorear columna Estado con fondo y negrita ===
    estado_colores = {
        "Rechazado": "FFA19A",
        "Com. Menores": "FFE5AD",
        "Com. Mayores": "DBB054",
        "Comentado": "F79646",
        "Enviado": "B1E1B9",
        "Sin Enviar": "FFFFAB",
        "Información": "FFFF46",
        "HOLD": "FF0909",
        "Aprobado": "00D25F"
    }

    for sheet in workbook.worksheets:
        col_estado_idx = None
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value == "Estado":
                col_estado_idx = idx
                break
        if col_estado_idx:
            for row in sheet.iter_rows(min_row=2):
                estado_val = row[col_estado_idx - 1].value
                if estado_val in estado_colores:
                    color_hex = estado_colores[estado_val]
                    row[col_estado_idx - 1].fill = PatternFill(start_color=color_hex,
                                                               end_color=color_hex,
                                                               fill_type="solid")
                    #row[col_estado_idx - 1].font = Font(bold=True) #Añadir que sea negrita


    # === Colorear columnas Responsable y Repsonsable solo texto y negrita ===
    responsables_colores = {
        "SS": "B22222",  # Azul marino oscuro
        "JM": "2E8B57",  # Verde esmeralda
        "JV": "007BA7",  # Azul petróleo
        "EC": "C71585",  # Rosa intenso (magenta oscuro)
        "ES": "6A0DAD",  # Púrpura fuerte
        "JP": "00509E",  # Azul profesional
        "AC": "4B0082",  # Índigo
        "CCH": "333333",  # Gris oscuro casi negro
        "LB": "1C86EE",  # Azul brillante
        "RM": "228B22",  # Verde bosque
        "RP": "1B365D",  # Rojo oscuro
        "EC/SS": "228B22" # Verde bosque
    }

    for sheet in workbook.worksheets:
        # Buscar índices de las columnas 'Responsable' y 'Repsonsable'
        col_indices = []
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value in ["Responsable", "Repsonsable"]:
                col_indices.append(idx)

        # Aplicar colores y negrita a ambas columnas si existen
        for col_idx in col_indices:
            for row in sheet.iter_rows(min_row=2):
                val = row[col_idx - 1].value
                if val in responsables_colores:
                    color_hex = responsables_colores[val]
                    row[col_idx - 1].font = Font(color=color_hex, bold=True)

    # === Colorear columna Crítico si contiene "Sí" (rojo y negrita) ===
    for sheet in workbook.worksheets:
        # Buscar índices de las columnas
        col_critico_idx = None
        col_dias_idx = None
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value == "Crítico":
                col_critico_idx = idx
            elif cell.value == "Días Devolución":
                col_dias_idx = idx

        # Aplicar formato independiente por columna
        if col_critico_idx:
            for row in sheet.iter_rows(min_row=2):
                if row[col_critico_idx - 1].value == "Sí":
                    row[col_critico_idx - 1].font = Font(color="FF0000", bold=True)

        if col_dias_idx:
            for row in sheet.iter_rows(min_row=2):
                valor = row[col_dias_idx - 1].value
                if valor is not None and isinstance(valor, (int, float)) and valor > 15:
                    row[col_dias_idx - 1].font = Font(color="FF0000", bold=True)

    # === Ajustar ancho de todas las columnas automáticamente ===
    for sheet in workbook.worksheets:
        for col_cells in sheet.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            sheet.column_dimensions[col_letter].width = max_length + 2

    workbook.save(archivo_excel)
    print("✅ Estilos aplicados a toda la tabla de datos.")