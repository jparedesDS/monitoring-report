# === GRÁFICO DE BARRAS APILADAS EN "STATUS GLOBAL" ===
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, Series

wb = load_workbook(output_path)

if "STATUS GLOBAL" in wb.sheetnames:
    ws = wb["STATUS GLOBAL"]

    # Buscar encabezados
    headers = [cell.value for cell in ws[1]]

    # Determinar las columnas que quieres graficar
    estado_cols = ["Aprobado", "Com. Mayores", "Com. Menores", "Enviado", "Rechazado", "Sin Enviar"]
    col_indices = [headers.index(col) + 1 for col in estado_cols]  # +1 porque Excel es 1-based

    # Rango de pedidos (columna A, desde fila 2 hasta el final)
    min_row = 2
    max_row = ws.max_row
    pedidos = Reference(ws, min_col=1, min_row=min_row, max_row=max_row)

    # Rango de datos (de B a G, según tus columnas seleccionadas)
    min_col = min(col_indices)
    max_col = max(col_indices)
    data = Reference(ws, min_col=min_col, max_col=max_col, min_row=1, max_row=max_row)

    # Crear gráfico de barras apiladas
    chart = BarChart()
    chart.type = "col"  # columnas verticales
    chart.title = "Estado por Pedido"
    chart.style = 12
    chart.grouping = "stacked"  # <- apiladas
    chart.overlap = 100
    chart.y_axis.title = "Nº Documentos"
    chart.x_axis.title = "Nº Pedido"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(pedidos)

    chart.height = 16
    chart.width = 29

    # Insertar gráfico en la hoja
    ws.add_chart(chart, "J3")

    wb.save(output_path)
    print("✅ Gráfico de barras apiladas añadido correctamente en la hoja de datos: 'STATUS GLOBAL'.")
else:
    print("⚠️ ERROR: No se encontró la hoja 'STATUS GLOBAL'.")